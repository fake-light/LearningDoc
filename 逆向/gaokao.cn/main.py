import requests
import hashlib
import hmac
import base64
import urllib.parse
from hashlib import md5
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from School import School


def w(encrypt_str):
    # 固定密钥，根据你的实际情况可能需要修改
    SIGN = "D23ABC@#56"

    # URL解码字符串
    decoded_str = urllib.parse.unquote(encrypt_str)

    # 使用HMAC-SHA1加密
    hmac_sha1 = hmac.new(SIGN.encode('utf-8'), decoded_str.encode('utf-8'), hashlib.sha1)

    # Base64编码
    base64_str = base64.b64encode(hmac_sha1.digest()).decode('utf-8')

    # MD5哈希
    md5_hash = md5(base64_str.encode('utf-8')).hexdigest()

    return md5_hash


def get_school_info(page, size, score):
    # 使用示例
    encrypted_str = f'api.zjzw.cn/web/api/?big_min={score + 15}&keyword=&local_batch_id=&local_province_id=52&local_type_id=2074&nature=36000&page={page}&province_id=&size={size}&small_min={score - 20}&type=&uri=apidata/api/gk/score/province&year=2024&zslx=0'  # 支持URL编码的输入
    signsafe = w(encrypted_str)
    print(f'signsafe =====>  {signsafe}')

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        'Origin': 'https://www.gaokao.cn',
        'Pragma': 'no-cache',
        'Referer': 'https://www.gaokao.cn/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    json_data = {
        'big_min': score + 15,
        'keyword': '',
        'local_batch_id': '',
        'local_province_id': '52',
        'local_type_id': '2074',
        'nature': '36000',
        'page': page,
        'province_id': '',
        'signsafe': signsafe,
        'size': size,
        'small_min': score - 20,
        'type': '',
        'uri': 'apidata/api/gk/score/province',
        'year': '2024',
        'zslx': '0',
    }

    response = requests.post(
        f'https://{encrypted_str}&signsafe={signsafe}',
        headers=headers,
        json=json_data,
    )
    return response.json()


def schools_to_excel_with_wrap(schools, filename):
    excel_data = []
    for school in schools:
        # 将专业信息格式化为字符串
        majors_info = "\n".join([
            f"{major.spname}: 录取人数{major.lq_num}, 最低分{major.min_score}, 最低位次{major.min_section}, 与分数线相差{major.diff}, 最低分相差{major.min_score - score if major.min_score - score > 0 else 0}"
            for major in school.majors
        ])
        excel_data.append({
            '学校名称': school.name,
            '省份': school.province,
            '城市': school.city_name,
            '符合条件专业信息': majors_info
        })

    df = pd.DataFrame(excel_data)
    df.to_excel(filename, index=False)

    # 使用openpyxl进行格式设置
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename)
        ws = wb.active

        # 设置专业信息列自动换行和垂直对齐
        col_letter = get_column_letter(df.columns.get_loc('符合条件专业信息') + 1)  # 获取列字母(如'D')

        # 设置列宽
        ws.column_dimensions[col_letter].width = 100  # 设置合适的列宽

        # 设置行高和自动换行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=df.columns.get_loc('符合条件专业信息') + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')  # 自动换行+顶部对齐
                # 根据换行数调整行高(每行按20像素估算，每个换行符增加一行)
                line_count = cell.value.count('\n') + 1
                ws.row_dimensions[cell.row].height = 15 * line_count  # 设置行高

        wb.save(filename)
        print(f"数据已保存到 {filename} (已设置自动换行)")
    except ImportError:
        print(f"数据已保存到 {filename} (但未设置格式，请安装openpyxl: pip install openpyxl)")
    except Exception as e:
        print(f"数据已保存到 {filename} (但格式设置失败: {str(e)})")


page = 1
size = 20
score = 480
local_province = 52
type_batch = '2074_14_0'
have_next_page = True

schools = []
while have_next_page:
    data = get_school_info(page, size, score)
    have_next_page = data['data']['numFound'] > page * size
    page += 1
    try:
        for item in data['data']['item']:
            majors = School.getMajor(item['school_id'], local_province, type_batch, score)
            if len(majors) == 0:
                continue
            school = School(item['school_id'], item['name'], item['province_name'], item['city_name'],
                            item['county_name'],
                            item['f211'],
                            item['f985'], item['nature_name'], majors)
            schools.append(school)
    except Exception as e:
        print(f'爬取学校出现问题了,错误信息: {e}')
        pass


schools_to_excel_with_wrap(schools, filename='school.xlsx')
